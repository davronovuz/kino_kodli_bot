import io
from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery, ContentType, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from loguru import logger

from filters.admin_filter import IsAdmin
from database.repositories import (
    MovieRepository, StatsRepository, UserRepository,
    CollectionRepository, AdvertisementRepository,
    DailyMovieRepository, MovieRequestRepository,
    AdminLogRepository,
)
from keyboards.inline import admin_menu_kb, confirm_kb, cancel_kb
from config import config

router = Router()
router.message.filter(IsAdmin())


# ============== FSM STATES ==============

class EditMovieStates(StatesGroup):
    waiting_code = State()
    waiting_field = State()
    waiting_value = State()


class AdStates(StatesGroup):
    waiting_text = State()
    waiting_photo = State()
    waiting_url = State()
    waiting_button = State()
    waiting_frequency = State()


class CollectionStates(StatesGroup):
    waiting_name = State()
    waiting_movie_code = State()


class ReplyRequestStates(StatesGroup):
    waiting_reply = State()


# ============== KINO TAHRIRLASH ==============

@router.message(F.text == "✏️ Kino tahrirlash")
async def edit_movie_start(message: Message, state: FSMContext):
    await state.set_state(EditMovieStates.waiting_code)
    await message.answer(
        "✏️ <b>Kino tahrirlash</b>\n\nKino kodini kiriting:",
        parse_mode="HTML", reply_markup=cancel_kb(),
    )


@router.message(EditMovieStates.waiting_code)
async def edit_movie_code(message: Message, state: FSMContext, session: AsyncSession):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=admin_menu_kb())
        return

    try:
        code = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return

    movie = await MovieRepository.get_by_code(session, code)
    if not movie:
        await message.answer("❌ Kino topilmadi!")
        return

    await state.update_data(edit_movie_id=movie.id, edit_movie_code=code)
    await state.set_state(EditMovieStates.waiting_field)

    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton

    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="📝 Nom", callback_data="efield:title"),
        InlineKeyboardButton(text="📅 Yil", callback_data="efield:year"),
    )
    builder.row(
        InlineKeyboardButton(text="📺 Sifat", callback_data="efield:quality"),
        InlineKeyboardButton(text="🌐 Til", callback_data="efield:language"),
    )
    builder.row(
        InlineKeyboardButton(text="📝 Tavsif", callback_data="efield:description"),
        InlineKeyboardButton(text="🔢 Kod", callback_data="efield:code"),
    )
    protect_text = "🔒 Himoya: ON" if movie.is_protected else "🔓 Himoya: OFF"
    builder.row(InlineKeyboardButton(text=protect_text, callback_data="efield:toggle_protect"))
    builder.row(InlineKeyboardButton(text="❌ Bekor", callback_data="efield:cancel"))

    await message.answer(
        f"✏️ <b>[{code}] {movie.title}</b>\n\nNimani o'zgartirmoqchisiz?",
        parse_mode="HTML", reply_markup=builder.as_markup(),
    )


@router.callback_query(EditMovieStates.waiting_field, F.data.startswith("efield:"))
async def edit_field_select(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    field = callback.data.split(":")[1]

    if field == "cancel":
        await state.clear()
        await callback.message.edit_text("❌ Bekor qilindi.")
        await callback.answer()
        return

    if field == "toggle_protect":
        data = await state.get_data()
        movie_id = data.get("edit_movie_id")
        movie = await MovieRepository.get_by_id(session, movie_id)
        if movie:
            new_val = not (movie.is_protected or False)
            await MovieRepository.update_movie(session, movie_id, is_protected=new_val)
            status = "🔒 Himoyalangan" if new_val else "🔓 Oddiy"
            await AdminLogRepository.log(
                session, callback.from_user.id,
                "movie_edit", f"[{data.get('edit_movie_code')}] is_protected={new_val}"
            )
            await callback.message.edit_text(f"✅ Kino himoyasi: {status}")
            await callback.message.answer("Admin menyu:", reply_markup=admin_menu_kb())
        await state.clear()
        await callback.answer()
        return

    field_names = {
        "title": "📝 Yangi nomni kiriting:",
        "year": "📅 Yangi yilni kiriting:",
        "quality": "📺 Yangi sifatni kiriting (720p, 1080p...):",
        "language": "🌐 Yangi tilni kiriting:",
        "description": "📝 Yangi tavsifni kiriting:",
        "code": "🔢 Yangi kodni kiriting:",
    }

    await state.update_data(edit_field=field)
    await state.set_state(EditMovieStates.waiting_value)
    await callback.message.edit_text(field_names.get(field, "Yangi qiymatni kiriting:"))
    await callback.answer()


@router.message(EditMovieStates.waiting_value)
async def edit_field_value(message: Message, state: FSMContext, session: AsyncSession):
    data = await state.get_data()
    movie_id = data.get("edit_movie_id")
    field = data.get("edit_field")

    if not movie_id or not field:
        await state.clear()
        return

    value = message.text.strip()

    try:
        if field == "year":
            value = int(value)
        elif field == "code":
            value = int(value)
            existing = await MovieRepository.get_by_code(session, value)
            if existing and existing.id != movie_id:
                await message.answer("❌ Bu kod band!")
                return

        kwargs = {field: value}
        movie = await MovieRepository.update_movie(session, movie_id, **kwargs)

        await AdminLogRepository.log(
            session, message.from_user.id,
            "movie_edit", f"[{data.get('edit_movie_code')}] {field}={value}"
        )

        from services.cache_service import CacheService
        await CacheService.invalidate_movie(data.get("edit_movie_code", 0))

        await message.answer(
            f"✅ Yangilandi!\n\n{field}: <b>{value}</b>",
            parse_mode="HTML", reply_markup=admin_menu_kb(),
        )
    except Exception as e:
        await message.answer(f"❌ Xato: {str(e)[:200]}")

    await state.clear()


# ============== EXCEL EXPORT ==============

@router.message(F.text == "📥 Excel export")
async def excel_export(message: Message, session: AsyncSession):
    progress = await message.answer("⏳ Excel tayyorlanmoqda...")

    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kinolar"

        headers = ["Kod", "Nom", "Yil", "Sifat", "Til", "Ko'rishlar", "Faol", "Qo'shilgan"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
            ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

        movies, total = await MovieRepository.get_all_movies(session, limit=10000, active_only=False)

        for row, movie in enumerate(movies, 2):
            ws.cell(row=row, column=1, value=movie.code)
            ws.cell(row=row, column=2, value=movie.title)
            ws.cell(row=row, column=3, value=movie.year)
            ws.cell(row=row, column=4, value=movie.quality)
            ws.cell(row=row, column=5, value=movie.language)
            ws.cell(row=row, column=6, value=movie.view_count)
            ws.cell(row=row, column=7, value="Ha" if movie.is_active else "Yo'q")
            ws.cell(row=row, column=8, value=movie.created_at.strftime("%d.%m.%Y") if movie.created_at else "")

        # Auto width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file = BufferedInputFile(buffer.read(), filename=f"kinolar_{total}.xlsx")
        await message.answer_document(file, caption=f"📊 Jami: {total} ta kino")
        await progress.delete()

    except Exception as e:
        await progress.edit_text(f"❌ Xato: {str(e)[:200]}")


# ============== REKLAMA BOSHQARUVI ==============

@router.message(F.text == "📣 Reklama")
async def manage_ads(message: Message, session: AsyncSession):
    ads = await AdvertisementRepository.get_all(session)

    text = "📣 <b>Reklama boshqaruvi</b>\n\n"
    if ads:
        for ad in ads[:5]:
            status = "✅" if ad.is_active else "❌"
            text += f"{status} ID:{ad.id} — {(ad.text or 'Rasm')[:30]}... (ko'rildi: {ad.view_count})\n"
    else:
        text += "Hozircha reklama yo'q.\n"

    text += "\n/newad — Yangi reklama\n/stopad — Reklamani o'chirish"
    await message.answer(text, parse_mode="HTML")


@router.message(F.text == "/newad")
async def new_ad_start(message: Message, state: FSMContext):
    await state.set_state(AdStates.waiting_text)
    await message.answer(
        "📣 <b>Yangi reklama</b>\n\nReklama matnini kiriting:",
        parse_mode="HTML", reply_markup=cancel_kb(),
    )


@router.message(AdStates.waiting_text)
async def ad_text(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=admin_menu_kb())
        return

    await state.update_data(ad_text=message.text)
    await state.set_state(AdStates.waiting_photo)

    from keyboards.inline import skip_kb
    await message.answer("🖼 Reklama rasmini yuboring (yoki o'tkazib yuboring):", reply_markup=skip_kb())


@router.message(AdStates.waiting_photo, F.content_type == ContentType.PHOTO)
async def ad_photo(message: Message, state: FSMContext):
    await state.update_data(ad_photo=message.photo[-1].file_id)
    await state.set_state(AdStates.waiting_url)

    from keyboards.inline import skip_kb
    await message.answer("🔗 Tugma URL kiriting (masalan https://t.me/kanal):", reply_markup=skip_kb())


@router.message(AdStates.waiting_photo, F.text == "⏭ O'tkazib yuborish")
async def ad_photo_skip(message: Message, state: FSMContext):
    await state.update_data(ad_photo=None)
    await state.set_state(AdStates.waiting_url)

    from keyboards.inline import skip_kb
    await message.answer("🔗 Tugma URL kiriting:", reply_markup=skip_kb())


@router.message(AdStates.waiting_url)
async def ad_url(message: Message, state: FSMContext):
    if message.text == "⏭ O'tkazib yuborish":
        await state.update_data(ad_url=None, ad_button=None)
        await state.set_state(AdStates.waiting_frequency)
        await message.answer("🔢 Har nechta kino ko'rishda reklama chiqsin? (masalan 5):")
        return

    await state.update_data(ad_url=message.text.strip())
    await state.set_state(AdStates.waiting_button)
    await message.answer("📝 Tugma matni (masalan 'Kanalga o'tish'):")


@router.message(AdStates.waiting_button)
async def ad_button(message: Message, state: FSMContext):
    await state.update_data(ad_button=message.text.strip())
    await state.set_state(AdStates.waiting_frequency)
    await message.answer("🔢 Har nechta kino ko'rishda reklama chiqsin? (masalan 5):")


@router.message(AdStates.waiting_frequency)
async def ad_frequency(message: Message, state: FSMContext, session: AsyncSession):
    try:
        freq = int(message.text.strip())
    except ValueError:
        await message.answer("Raqam kiriting!")
        return

    data = await state.get_data()

    # Eski reklamalarni o'chirish
    await AdvertisementRepository.deactivate_all(session)

    ad = await AdvertisementRepository.create(
        session,
        text=data.get("ad_text"),
        photo_file_id=data.get("ad_photo"),
        url=data.get("ad_url"),
        button_text=data.get("ad_button"),
        show_every=freq,
        is_active=True,
    )

    await state.clear()
    await message.answer(
        f"✅ Reklama yaratildi!\nHar {freq} ta kino ko'rishda chiqadi.",
        reply_markup=admin_menu_kb(),
    )


@router.message(F.text == "/stopad")
async def stop_ads(message: Message, session: AsyncSession):
    await AdvertisementRepository.deactivate_all(session)
    await message.answer("✅ Barcha reklamalar o'chirildi.")


# ============== KINO TO'PLAM ==============

@router.message(F.text == "📂 To'plamlar")
async def manage_collections_admin(message: Message, session: AsyncSession):
    cols = await CollectionRepository.get_all(session)

    text = "📂 <b>Kino to'plamlar</b>\n\n"
    if cols:
        for col in cols:
            movies = await CollectionRepository.get_movies(session, col.id)
            text += f"{col.emoji} <b>{col.name}</b> — {len(movies)} ta kino\n"
    else:
        text += "Hozircha to'plam yo'q.\n"

    text += "\n/newcol Nom — Yangi to'plam yaratish\n/addtocol TO'PLAM_ID KINO_KODI — Kino qo'shish"
    await message.answer(text, parse_mode="HTML")


@router.message(F.text.startswith("/newcol"))
async def new_collection(message: Message, session: AsyncSession):
    name = message.text.replace("/newcol", "").strip()
    if not name:
        await message.answer("Foydalanish: /newcol To'plam nomi")
        return

    col = await CollectionRepository.create(session, name=name)
    await message.answer(f"✅ To'plam yaratildi: <b>{name}</b> (ID: {col.id})", parse_mode="HTML")


@router.message(F.text.startswith("/addtocol"))
async def add_to_collection(message: Message, session: AsyncSession):
    args = message.text.split()
    if len(args) < 3:
        await message.answer("Foydalanish: /addtocol TO'PLAM_ID KINO_KODI")
        return

    try:
        col_id = int(args[1])
        movie_code = int(args[2])
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return

    col = await CollectionRepository.get_by_id(session, col_id)
    if not col:
        await message.answer("❌ To'plam topilmadi!")
        return

    movie = await MovieRepository.get_by_code(session, movie_code)
    if not movie:
        await message.answer("❌ Kino topilmadi!")
        return

    try:
        await CollectionRepository.add_movie(session, col_id, movie.id)
        await message.answer(f"✅ <b>{movie.title}</b> → <b>{col.name}</b> ga qo'shildi!", parse_mode="HTML")
    except Exception:
        await message.answer("⚠️ Bu kino allaqachon to'plamda.")


# ============== KINO SO'ROVLARIGA JAVOB ==============

@router.message(F.text.startswith("/reply_"))
async def reply_to_request(message: Message, session: AsyncSession, bot: Bot):
    try:
        parts = message.text.split(" ", 1)
        req_id = int(parts[0].replace("/reply_", ""))
        reply_text = parts[1] if len(parts) > 1 else "Javob berildi"
    except (ValueError, IndexError):
        await message.answer("Foydalanish: /reply_ID matn")
        return

    req = await MovieRequestRepository.get_by_id(session, req_id)
    if not req:
        await message.answer("❌ So'rov topilmadi!")
        return

    await MovieRequestRepository.reply(session, req_id, reply_text)

    try:
        await bot.send_message(
            req.user_id,
            f"📩 <b>So'rovingizga javob!</b>\n\n"
            f"🎬 So'rov: <i>{req.request_text}</i>\n"
            f"💬 Javob: {reply_text}",
            parse_mode="HTML",
        )
    except Exception:
        pass

    await message.answer("✅ Javob yuborildi!")


# ============== KUNLIK KINO TANLASH ==============

@router.message(F.text.startswith("/setdaily"))
async def set_daily_movie(message: Message, session: AsyncSession):
    args = message.text.split()
    if len(args) < 2:
        await message.answer("Foydalanish: /setdaily KINO_KODI")
        return

    try:
        code = int(args[1])
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return

    movie = await MovieRepository.get_by_code(session, code)
    if not movie:
        await message.answer("❌ Kino topilmadi!")
        return

    await DailyMovieRepository.set_today(session, movie.id)
    await message.answer(f"✅ Bugungi kino: <b>{movie.title}</b>", parse_mode="HTML")


# ============== KINO O'CHIRISH ==============

class DeleteMovieStates(StatesGroup):
    waiting_code = State()
    waiting_confirm = State()




@router.message(F.text == "🗑 Kino o'chirish")
async def delete_movie_start(message: Message, state: FSMContext):
    await state.set_state(DeleteMovieStates.waiting_code)
    await message.answer(
        "🗑 <b>Kino o'chirish</b>\n\nKino kodini kiriting:",
        parse_mode="HTML", reply_markup=cancel_kb(),
    )


@router.message(DeleteMovieStates.waiting_code)
async def delete_movie_code(message: Message, state: FSMContext, session: AsyncSession):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("Bekor qilindi.", reply_markup=admin_menu_kb())
        return

    try:
        code = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return

    movie = await MovieRepository.get_by_code(session, code)
    if not movie:
        await message.answer("❌ Kino topilmadi!")
        return

    await state.update_data(delete_movie_id=movie.id, delete_movie_code=code, delete_movie_title=movie.title)
    await state.set_state(DeleteMovieStates.waiting_confirm)

    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton

    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="✅ Ha, o'chirish", callback_data="confirm_delete:yes"),
        InlineKeyboardButton(text="❌ Yo'q", callback_data="confirm_delete:no"),
    )

    await message.answer(
        f"🗑 Rostdan o'chirilsinmi?\n\n"
        f"🎬 <b>[{code}] {movie.title}</b>\n"
        f"👁 Ko'rishlar: {movie.view_count}",
        parse_mode="HTML", reply_markup=builder.as_markup(),
    )


@router.callback_query(DeleteMovieStates.waiting_confirm, F.data.startswith("confirm_delete:"))
async def delete_movie_confirm(callback: CallbackQuery, state: FSMContext, session: AsyncSession):
    answer = callback.data.split(":")[1]

    if answer == "no":
        await state.clear()
        await callback.message.edit_text("❌ Bekor qilindi.")
        await callback.answer()
        return

    data = await state.get_data()
    movie_id = data.get("delete_movie_id")
    code = data.get("delete_movie_code")
    title = data.get("delete_movie_title")

    try:
        await MovieRepository.delete_movie(session, movie_id)
        await AdminLogRepository.log(
            session, callback.from_user.id,
            "movie_delete", f"[{code}] {title}"
        )
        from services.cache_service import CacheService
        await CacheService.invalidate_movie(code)

        await callback.message.edit_text(
            f"✅ O'chirildi!\n\n🎬 <b>[{code}] {title}</b>",
            parse_mode="HTML",
        )
    except Exception as e:
        await callback.message.edit_text(f"❌ Xato: {str(e)[:200]}")

    await state.clear()
    await callback.answer()


# ============== AUDIT LOG ==============

@router.message(F.text == "📋 Audit log")
async def show_audit_log(message: Message, session: AsyncSession):
    logs = await AdminLogRepository.get_recent(session, limit=15)

    if not logs:
        await message.answer("📋 <b>Audit log bo'sh.</b>", parse_mode="HTML")
        return

    text = "📋 <b>Oxirgi admin harakatlari:</b>\n\n"
    for log in logs:
        date_str = log.created_at.strftime("%d.%m %H:%M")
        detail = f" — {log.detail[:50]}" if log.detail else ""
        text += f"🕐 {date_str} | 👤 {log.admin_id}\n   {log.action}{detail}\n\n"

    await message.answer(text, parse_mode="HTML")