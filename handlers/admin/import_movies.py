import asyncio
from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery, ContentType
from aiogram.fsm.context import FSMContext
from sqlalchemy.ext.asyncio import AsyncSession
from loguru import logger

from filters.admin_filter import IsAdmin
from database.repositories import MovieRepository
from states.admin_states import ImportStates
from keyboards.inline import import_method_kb, cancel_kb, admin_menu_kb
from services.cache_service import CacheService
from config import config

router = Router()
router.message.filter(IsAdmin())


@router.message(F.text == "📥 Import kinolar")
async def start_import(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        "📥 <b>Kinolarni import qilish</b>\n\n"
        "Usulni tanlang:",
        reply_markup=import_method_kb(),
        parse_mode="HTML",
    )


# ============== FORWARD IMPORT ==============

@router.callback_query(F.data == "import:forward")
async def import_forward_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ImportStates.waiting_forward)
    await state.update_data(imported_count=0, failed_count=0, skipped_count=0)
    await callback.message.edit_text(
        "📤 <b>Forward import</b>\n\n"
        "Kinolarni botga forward qiling.\n"
        "Har bir video/dokumentga avtomatik kod beriladi.\n\n"
        "⚠️ <b>Muhim:</b> Telegram limitlari tufayli biroz sekin forward qiling.\n\n"
        "Tugatgach «❌ Bekor qilish» bosing yoki /done yozing.",
        parse_mode="HTML",
    )
    await callback.message.answer("Kinolarni forward qiling:", reply_markup=cancel_kb())
    await callback.answer()


@router.message(
    ImportStates.waiting_forward,
    F.content_type.in_({ContentType.VIDEO, ContentType.DOCUMENT}),
)
async def import_forward_receive(message: Message, state: FSMContext, session: AsyncSession):
    data = await state.get_data()
    imported = data.get("imported_count", 0)
    failed = data.get("failed_count", 0)
    skipped = data.get("skipped_count", 0)

    if message.video:
        file_id = message.video.file_id
        file_unique_id = message.video.file_unique_id
        file_type = "video"
        duration = message.video.duration
        file_size = message.video.file_size
    elif message.document:
        file_id = message.document.file_id
        file_unique_id = message.document.file_unique_id
        file_type = "document"
        duration = None
        file_size = message.document.file_size
    else:
        return

    # Check duplicate
    existing = await MovieRepository.get_by_file_unique_id(session, file_unique_id)
    if existing:
        skipped += 1
        await state.update_data(skipped_count=skipped)
        await message.reply(
            f"⏭ O'tkazildi (allaqachon bor): kod <code>{existing.code}</code>",
            parse_mode="HTML",
        )
        return

    # Extract title from caption or filename
    title = "Nomsiz kino"
    if message.caption:
        title = message.caption.strip()[:500]
    elif message.document and message.document.file_name:
        # Remove extension
        fname = message.document.file_name
        title = fname.rsplit(".", 1)[0] if "." in fname else fname

    try:
        code = await MovieRepository.get_next_code(session)
        movie = await MovieRepository.create(
            session,
            code=code,
            title=title,
            file_id=file_id,
            file_unique_id=file_unique_id,
            file_type=file_type,
            duration=duration,
            file_size=file_size,
            caption=message.caption,
            added_by=message.from_user.id,
        )
        imported += 1
        await state.update_data(imported_count=imported)
        await message.reply(
            f"✅ Qo'shildi! Kod: <code>{movie.code}</code> | {title[:50]}",
            parse_mode="HTML",
        )
    except Exception as e:
        failed += 1
        await state.update_data(failed_count=failed)
        logger.error(f"Import error: {e}")
        await message.reply(f"❌ Xato: {str(e)[:100]}")


@router.message(ImportStates.waiting_forward, F.text == "❌ Bekor qilish")
@router.message(ImportStates.waiting_forward, F.text == "/done")
async def import_forward_done(message: Message, state: FSMContext):
    data = await state.get_data()
    imported = data.get("imported_count", 0)
    failed = data.get("failed_count", 0)
    skipped = data.get("skipped_count", 0)

    await state.clear()
    await CacheService.invalidate_all_movies()

    await message.answer(
        f"✅ <b>Import yakunlandi!</b>\n\n"
        f"✅ Qo'shildi: {imported}\n"
        f"⏭ O'tkazildi: {skipped}\n"
        f"❌ Xato: {failed}\n"
        f"📊 Jami: {imported + failed + skipped}",
        reply_markup=admin_menu_kb(),
        parse_mode="HTML",
    )


# ============== GROUP IMPORT ==============

@router.callback_query(F.data == "import:group")
async def import_group_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ImportStates.waiting_group_id)
    await callback.message.edit_text(
        "📤 <b>Guruhdan import</b>\n\n"
        "Botni guruhga admin qilib qo'shing, keyin guruh ID sini yuboring.\n\n"
        "Guruh ID sini bilish uchun guruhga /id buyrug'ini yuboring yoki\n"
        "botni guruhga qo'shib, guruhda biror xabar yozing.\n\n"
        "Masalan: <code>-1001234567890</code>",
        parse_mode="HTML",
    )
    await callback.message.answer("Guruh ID sini yuboring:", reply_markup=cancel_kb())
    await callback.answer()


@router.message(ImportStates.waiting_group_id)
async def import_group_receive(
    message: Message, state: FSMContext, session: AsyncSession, bot: Bot
):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("❌ Bekor qilindi.", reply_markup=admin_menu_kb())
        return

    try:
        group_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Noto'g'ri ID format! Raqam kiriting.")
        return

    # Verify bot is member
    try:
        chat = await bot.get_chat(group_id)
        member = await bot.get_chat_member(group_id, bot.id)
        if member.status not in ("administrator", "creator"):
            await message.answer(
                "❌ Bot bu guruhda admin emas!\n"
                "Botni guruhga admin qiling va qaytadan urinib ko'ring."
            )
            return
    except Exception as e:
        await message.answer(f"❌ Guruhga kirib bo'lmadi: {str(e)[:200]}")
        return

    await state.set_state(ImportStates.processing)

    progress_msg = await message.answer(
        f"📥 <b>Import boshlandi...</b>\n"
        f"📢 Guruh: {chat.title}\n"
        f"⏳ Bu biroz vaqt olishi mumkin...",
        parse_mode="HTML",
    )

    # Note: Telegram Bot API doesn't support message history access
    # Users need to forward messages from the group
    await message.answer(
        "⚠️ <b>Muhim!</b>\n\n"
        "Telegram Bot API guruh tarixini o'qiy olmaydi.\n\n"
        "🔄 <b>Muqobil usul:</b>\n"
        "1. Guruhdan kinolarni tanlang\n"
        "2. Botga forward qiling\n"
        "3. Bot avtomatik kod beradi\n\n"
        "Yoki Telegram Client API (Pyrogram/Telethon) dan foydalaning.\n\n"
        "Hozir «Forward import» usulini tanlang.",
        parse_mode="HTML",
    )

    await state.clear()
    await message.answer("Admin menyu:", reply_markup=admin_menu_kb())


# ============== EXCEL IMPORT ==============

@router.callback_query(F.data == "import:excel")
async def import_excel_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ImportStates.waiting_file)
    await callback.message.edit_text(
        "📄 <b>Excel/CSV import</b>\n\n"
        "Excel yoki CSV faylni yuboring.\n\n"
        "📋 <b>Format:</b>\n"
        "| code | title | year | quality | language |\n"
        "|------|-------|------|---------|----------|\n"
        "| 1 | Venom | 2024 | 1080p | O'zbek tilida |\n\n"
        "⚠️ Bu usul faqat ma'lumotlarni bazaga yozadi.\n"
        "Video fayllarni alohida qo'shish kerak bo'ladi.",
        parse_mode="HTML",
    )
    await callback.message.answer("Faylni yuboring:", reply_markup=cancel_kb())
    await callback.answer()


@router.message(ImportStates.waiting_file, F.content_type == ContentType.DOCUMENT)
async def import_excel_receive(message: Message, state: FSMContext, session: AsyncSession, bot: Bot):
    doc = message.document
    if not doc.file_name:
        await message.answer("❌ Fayl nomi topilmadi!")
        return

    ext = doc.file_name.rsplit(".", 1)[-1].lower() if "." in doc.file_name else ""
    if ext not in ("xlsx", "xls", "csv"):
        await message.answer("❌ Faqat .xlsx, .xls yoki .csv fayllar qabul qilinadi!")
        return

    progress = await message.answer("⏳ Fayl yuklanmoqda...")

    try:
        file = await bot.download(doc)
        import io

        if ext == "csv":
            import csv
            content = file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [str(cell.value).lower().strip() for cell in ws[1] if cell.value]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                if any(row_dict.values()):
                    rows.append(row_dict)

        if not rows:
            await progress.edit_text("❌ Fayl bo'sh yoki format noto'g'ri!")
            await state.clear()
            return

        imported = 0
        failed = 0
        skipped = 0

        for row in rows:
            try:
                code = row.get("code")
                title = row.get("title") or row.get("nom") or row.get("name")

                if not title:
                    failed += 1
                    continue

                if code:
                    code = int(code)
                    existing = await MovieRepository.get_by_code(session, code)
                    if existing:
                        skipped += 1
                        continue
                else:
                    code = await MovieRepository.get_next_code(session)

                await MovieRepository.create(
                    session,
                    code=code,
                    title=str(title).strip(),
                    year=int(row.get("year") or row.get("yil") or 0) or None,
                    quality=str(row.get("quality") or row.get("sifat") or "").strip() or None,
                    language=str(row.get("language") or row.get("til") or "").strip() or None,
                    file_id="PLACEHOLDER_" + str(code),
                    added_by=message.from_user.id,
                )
                imported += 1

            except Exception as e:
                failed += 1
                logger.error(f"Excel import row error: {e}")

        await progress.edit_text(
            f"✅ <b>Excel import yakunlandi!</b>\n\n"
            f"✅ Qo'shildi: {imported}\n"
            f"⏭ O'tkazildi: {skipped}\n"
            f"❌ Xato: {failed}\n"
            f"📊 Jami: {len(rows)} qator",
            parse_mode="HTML",
        )
        await CacheService.invalidate_all_movies()

    except Exception as e:
        logger.error(f"Excel import error: {e}")
        await progress.edit_text(f"❌ Xatolik: {str(e)[:300]}")

    await state.clear()
    await message.answer("Admin menyu:", reply_markup=admin_menu_kb())


@router.message(ImportStates.waiting_file, F.text == "❌ Bekor qilish")
async def import_excel_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=admin_menu_kb())


@router.callback_query(F.data == "import:cancel")
async def import_cancel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("❌ Bekor qilindi.")
    await callback.message.answer("Admin menyu:", reply_markup=admin_menu_kb())
    await callback.answer()
